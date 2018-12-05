import openpyxl
import os
class Operate_Excel:

    def __init__(self, filepath, filename):
        self.__file = filepath + '/' + filename
        if os.path.exists(self.__file):
            print('Open excel:', self.__file)
        else:
            raise ('excel file is not exit')

    def getExcelobject(self):
        """返回excel对象"""
        try:
            return openpyxl.load_workbook(self.__file)

        except Exception as e:
            raise e

    def getSheetobjcet(self,sheetname):
        try:
            return openpyxl.load_workbook(self.__file)[sheetname]
        except Exception as e:
            raise e

    def getCellogject(self, sheetobject, cellname, row=None, column=None):
        try:
            if cellname:
                return sheetobject[cellname]
            elif not cellname and row and column:
                return sheetobject.cell(row, column)
            else:
                raise "Cell's row or column is wrong"
        except Exception as e:
            raise e

    def getCellvalue(self, sheetobject, cellname, row=None, column=None):
        try:
            if cellname:
                return sheetobject[cellname].value
            elif not cellname and row and column:
                return sheetobject.cell(row, column).value
            else:
                raise "Cell's row or column is wrong"
        except Exception as e:
            raise e







print('hahah')
file = Operate_Excel('C:/workspace/api_auto_test/api_auto_test/data','api.xlsx')
print(file.getExcelobject())
print(file.getSheetobjcet('登录'))

